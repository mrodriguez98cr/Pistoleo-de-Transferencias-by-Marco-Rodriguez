import customtkinter as ctk
from tkinter import ttk, filedialog, messagebox
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

ctk.set_appearance_mode("light")
ctk.set_default_color_theme("blue")

# ── Paleta Material Design ───────────────────────
C_PRIMARY      = "#C62828"   # Red 800
C_PRIMARY_DARK = "#8E0000"   # Red 900
C_PRIMARY_LIGHT= "#FF5F52"   # Red 400
C_ON_PRIMARY   = "#FFFFFF"
C_BG           = "#EEEEEE"   # Grey 200 — fondo general
C_SURFACE      = "#FFFFFF"   # Cards y paneles
C_ON_SURFACE   = "#212121"   # Texto principal
C_SECONDARY    = "#757575"   # Texto secundario
C_DIVIDER      = "#BDBDBD"   # Separadores
C_GREEN        = "#2E7D32"
C_AMBER        = "#F57F17"
C_BLUE         = "#1565C0"
C_RED_TINT     = "#FFEBEE"


class MainView(ctk.CTk):

    def __init__(self, controller=None):
        super().__init__()
        self.controller = controller
        self.botones_transferencias = {}

        self.title("Pistoleo de Transferencias by Marco Rodriguez")
        self.state("zoomed")
        self.configure(fg_color=C_BG)

        self._apply_styles()
        self._build_ui()

    # ──────────────────────────────────────────────
    # Estilos ttk (Material-like)
    # ──────────────────────────────────────────────

    def _apply_styles(self):
        s = ttk.Style()
        s.theme_use("clam")

        s.configure("M.Treeview",
            background=C_SURFACE,
            fieldbackground=C_SURFACE,
            foreground=C_ON_SURFACE,
            rowheight=44,
            font=("Roboto", 12) if self._font_exists("Roboto") else ("Segoe UI", 12),
            borderwidth=0,
        )
        s.configure("M.Treeview.Heading",
            background=C_PRIMARY,
            foreground=C_ON_PRIMARY,
            font=("Segoe UI", 11, "bold"),
            relief="flat",
            padding=(12, 10),
        )
        s.map("M.Treeview",
            background=[("selected", C_RED_TINT)],
            foreground=[("selected", C_PRIMARY)],
        )

    def _font_exists(self, name):
        try:
            import tkinter.font as f
            return name in f.families()
        except Exception:
            return False

    # ──────────────────────────────────────────────
    # Build UI
    # ──────────────────────────────────────────────

    def _build_ui(self):
        self._build_appbar()
        self._build_stats()
        self._build_scanner()
        self._build_body()

    # ── App Bar (Material) ───────────────────────

    def _build_appbar(self):
        bar = ctk.CTkFrame(self, fg_color=C_PRIMARY, corner_radius=0, height=56)
        bar.pack(fill="x")
        bar.pack_propagate(False)

        ctk.CTkLabel(
            bar,
            text="Pistoleo de Transferencias by Marco Rodriguez",
            font=ctk.CTkFont("Segoe UI", 20, weight="bold"),
            text_color=C_ON_PRIMARY,
        ).pack(side="left", padx=20, pady=14)

        # Acciones en la app bar
        btn_cfg = dict(
            height=34,
            corner_radius=4,
            fg_color="transparent",
            hover_color=C_PRIMARY_DARK,
            text_color=C_ON_PRIMARY,
            border_width=1,
            border_color="#EF9A9A",
            font=ctk.CTkFont("Segoe UI", 12, weight="bold"),
        )

        actions = [
            ("Cargar UPC",      self._on_cargar_upc),
            ("Transferencia",   self._on_cargar_transferencia),
            ("Reporte",         self._on_exportar_reporte),
            ("Ordenar",         self._on_ordenar),
            ("Limpiar",         self._on_limpiar),
        ]

        for texto, cmd in reversed(actions):
            ctk.CTkButton(bar, text=texto, width=130, command=cmd, **btn_cfg).pack(
                side="right", padx=6, pady=10
            )

    # ── Stats Cards ──────────────────────────────

    def _build_stats(self):
        frame = ctk.CTkFrame(self, fg_color=C_BG, corner_radius=0)
        frame.pack(fill="x", padx=16, pady=12)

        self.stats = {}
        items = [
            ("Transferidas", "📦", C_BLUE),
            ("Pistoleadas",  "✔",  C_GREEN),
            ("Fuera",        "⚠",  C_AMBER),
            ("Diferencia",   "Δ",  C_PRIMARY),
        ]

        for nombre, icono, color in items:
            # Card con sombra simulada
            shadow = ctk.CTkFrame(frame, fg_color="#D5D5D5", corner_radius=8)
            shadow.pack(side="left", fill="x", expand=True, padx=6, pady=(2, 0))

            card = ctk.CTkFrame(shadow, fg_color=C_SURFACE, corner_radius=8)
            card.pack(fill="both", expand=True, padx=1, pady=(0, 2))

            # Franja de color arriba
            ctk.CTkFrame(card, fg_color=color, height=3, corner_radius=0).pack(fill="x")

            body = ctk.CTkFrame(card, fg_color="transparent")
            body.pack(fill="x", padx=16, pady=10)

            ctk.CTkLabel(
                body,
                text=f"{icono}  {nombre.upper()}",
                font=ctk.CTkFont("Segoe UI", 10, weight="bold"),
                text_color=C_SECONDARY,
            ).pack(anchor="w")

            valor = ctk.CTkLabel(
                body,
                text="0",
                font=ctk.CTkFont("Segoe UI", 34, weight="bold"),
                text_color=color,
            )
            valor.pack(anchor="w", pady=(2, 0))
            self.stats[nombre] = valor

    # ── Scanner Bar ──────────────────────────────

    def _build_scanner(self):
        card = ctk.CTkFrame(self, fg_color=C_SURFACE, corner_radius=8)
        card.pack(fill="x", padx=16, pady=(0, 12))

        inner = ctk.CTkFrame(card, fg_color="transparent")
        inner.pack(fill="x", padx=16, pady=12)

        ctk.CTkLabel(
            inner,
            text="UPC",
            font=ctk.CTkFont("Segoe UI", 11, weight="bold"),
            text_color=C_SECONDARY,
        ).pack(side="left", padx=(0, 10))

        self.entry_upc = ctk.CTkEntry(
            inner,
            width=360,
            height=40,
            corner_radius=4,
            fg_color=C_BG,
            border_color=C_PRIMARY,
            border_width=2,
            font=ctk.CTkFont("Segoe UI", 16, weight="bold"),
            text_color=C_ON_SURFACE,
            placeholder_text="Escanear o ingresar UPC...",
        )
        self.entry_upc.pack(side="left", padx=(0, 10))
        self.entry_upc.bind("<Return>", self._on_verificar_upc)
        self.entry_upc.focus()

        ctk.CTkButton(
            inner,
            text="VERIFICAR",
            width=120,
            height=40,
            corner_radius=4,
            fg_color=C_PRIMARY,
            hover_color=C_PRIMARY_DARK,
            text_color=C_ON_PRIMARY,
            font=ctk.CTkFont("Segoe UI", 13, weight="bold"),
            command=self._on_verificar_upc,
        ).pack(side="left")

        self.lbl_ultimo = ctk.CTkLabel(
            inner,
            text="— sin escaneos —",
            font=ctk.CTkFont("Segoe UI", 20),
            text_color=C_SECONDARY,
        )
        self.lbl_ultimo.pack(side="left", padx=20)

    # ── Body: izquierda + centro + derecha ───────

    def _build_body(self):
        body = ctk.CTkFrame(self, fg_color=C_BG, corner_radius=0)
        body.pack(fill="both", expand=True, padx=16, pady=(0, 16))

        self._panel_izquierdo(body)
        self._panel_derecho(body)
        self._panel_central(body)

    # ── Panel Izquierdo — Escaneados ─────────────

    def _panel_izquierdo(self, parent):
        card = self._card(parent, side="left", width=220)

        ctk.CTkLabel(
            card,
            text="ESCANEADOS",
            font=ctk.CTkFont("Segoe UI", 10, weight="bold"),
            text_color=C_SECONDARY,
        ).pack(padx=20, pady=(14, 4), anchor="w")

        ctk.CTkFrame(card, fg_color=C_DIVIDER, height=1).pack(fill="x", padx=12)

        self.text_upcs = ctk.CTkTextbox(
            card,
            fg_color=C_SURFACE,
            text_color=C_PRIMARY,
            font=ctk.CTkFont("Segoe UI", 34, weight="bold"),
            border_width=0,
            scrollbar_button_color=C_DIVIDER,
        )
        self.text_upcs.pack(fill="both", expand=True, padx=8, pady=8)

    # ── Panel Central — Tabla ────────────────────

    def _panel_central(self, parent):
        card = self._card(parent, side="left")

        ctk.CTkLabel(
            card,
            text="ARTÍCULOS",
            font=ctk.CTkFont("Segoe UI", 10, weight="bold"),
            text_color=C_SECONDARY,
        ).pack(padx=16, pady=(14, 4), anchor="w")

        ctk.CTkFrame(card, fg_color=C_DIVIDER, height=1).pack(fill="x", padx=12)

        tree_wrap = ctk.CTkFrame(card, fg_color="transparent")
        tree_wrap.pack(fill="both", expand=True, padx=8, pady=8)

        scroll = ttk.Scrollbar(tree_wrap, orient="vertical")
        scroll.pack(side="right", fill="y")

        self.tree = ttk.Treeview(
            tree_wrap,
            style="M.Treeview",
            columns=("codigo", "descripcion", "transferencia", "pistoleada"),
            show="headings",
            yscrollcommand=scroll.set,
        )
        scroll.config(command=self.tree.yview)
        self.tree.pack(fill="both", expand=True)

        for col, texto, ancho, anchor in [
            ("codigo",        "CÓDIGO",        161, "center"),
            ("descripcion",   "DESCRIPCIÓN",   400, "w"),
            ("transferencia", "TRANSFERENCIA", 140, "center"),
            ("pistoleada",    "PISTOLEADA",    140, "center"),
        ]:
            self.tree.heading(col, text=texto)
            self.tree.column(col, width=ancho, anchor=anchor)

        self.tree.tag_configure("verde",    background="#E8F5E9", foreground=C_GREEN)
        self.tree.tag_configure("rojo",     background="#FFEBEE", foreground=C_PRIMARY)
        self.tree.tag_configure("amarillo", background="#FFF8E1", foreground=C_AMBER)
        self.tree.tag_configure("azul",     background="#E3F2FD", foreground=C_BLUE)

    # ── Panel Derecho — Transferencias + Fuera ───

    def _panel_derecho(self, parent):
        card = self._card(parent, side="right", width=260)

        # Transferencias cargadas
        ctk.CTkLabel(
            card,
            text="TRANSFERENCIAS",
            font=ctk.CTkFont("Segoe UI", 10, weight="bold"),
            text_color=C_SECONDARY,
        ).pack(padx=16, pady=(14, 4), anchor="w")

        ctk.CTkFrame(card, fg_color=C_DIVIDER, height=1).pack(fill="x", padx=12)

        self.frame_transferencias = ctk.CTkScrollableFrame(
            card,
            fg_color=C_SURFACE,
            scrollbar_button_color=C_DIVIDER,
            height=180,
        )
        self.frame_transferencias.pack(fill="x", padx=8, pady=8)

        # Divider
        ctk.CTkFrame(card, fg_color=C_DIVIDER, height=1).pack(fill="x", padx=12, pady=(4, 0))

        # Fuera de transferencia
        ctk.CTkLabel(
            card,
            text="FUERA DE TRANSFERENCIA",
            font=ctk.CTkFont("Segoe UI", 10, weight="bold"),
            text_color=C_SECONDARY,
        ).pack(padx=16, pady=(12, 4), anchor="w")

        ctk.CTkFrame(card, fg_color=C_DIVIDER, height=1).pack(fill="x", padx=12)

        self.text_fuera = ctk.CTkTextbox(
            card,
            fg_color=C_SURFACE,
            text_color=C_SECONDARY,
            font=ctk.CTkFont("Segoe UI", 12),
            border_width=0,
            scrollbar_button_color=C_DIVIDER,
        )
        self.text_fuera.pack(fill="both", expand=True, padx=8, pady=8)

    # ── Helper: card con sombra ──────────────────

    def _card(self, parent, side="left", width=None, expand=True):
        """Crea una card Material con sombra simulada."""
        shadow = ctk.CTkFrame(parent, fg_color="#D5D5D5", corner_radius=8)
        kw = dict(side=side, fill="both", padx=6, pady=2)
        if width:
            kw["fill"] = "y"
            kw.pop("expand", None)
            shadow.configure(width=width)
            shadow.pack_propagate(False)
        else:
            kw["expand"] = expand
        shadow.pack(**kw)

        card = ctk.CTkFrame(shadow, fg_color=C_SURFACE, corner_radius=8)
        card.pack(fill="both", expand=True, padx=1, pady=(0, 2))
        return card

    # ──────────────────────────────────────────────
    # Eventos
    # ──────────────────────────────────────────────

    def _on_cargar_upc(self):
        ruta = filedialog.askopenfilename(filetypes=[("Excel", "*.xlsx")])
        if not ruta:
            return
        if self.controller:
            exito, msg = self.controller.cargar_upc(ruta)
            messagebox.showinfo("Éxito", msg) if exito else messagebox.showerror("Error", msg)

    def _on_cargar_transferencia(self):
        ruta = filedialog.askopenfilename(filetypes=[("Excel", "*.xlsx")])
        if not ruta:
            return
        nombre = Path(ruta).stem
        if self.controller:
            exito, msg, filas, nombre_unico = self.controller.cargar_transferencia(ruta, nombre)
            if not exito:
                messagebox.showerror("Error", msg)
                return
            self._refrescar_treeview(filas)
            self._actualizar_contadores()
        self._agregar_boton_transferencia(nombre_unico)

    def _on_verificar_upc(self, event=None):
        upc = self.entry_upc.get().strip()
        self.entry_upc.delete(0, "end")
        if not upc:
            return
        if self.controller:
            resultado = self.controller.escanear_upc(upc)
            if not resultado["encontrado"]:
                messagebox.showwarning("Alerta", resultado["mensaje"])
                return
            self.lbl_ultimo.configure(
                text=f"✔  {resultado['codigo']}  ·  {resultado['descripcion']}"
            )
            self.text_upcs.insert("end", resultado["codigo"] + "\n")
            self._actualizar_fila_treeview(resultado)
            if resultado.get("sobrante"):
                messagebox.showwarning("Sobrante", resultado["mensaje"])
            if not resultado["en_transferencia"]:
                self.text_fuera.insert("end", f"{resultado['codigo']} - {resultado['descripcion']}\n")
                messagebox.showwarning("Fuera", resultado["mensaje"])
            self._actualizar_contadores()
            self._actualizar_colores()
        else:
            self.text_upcs.insert("end", upc + "\n")
            self.lbl_ultimo.configure(text=f"✔  {upc}")

    def _on_limpiar(self):
        if not messagebox.askyesno("Limpiar", "¿Limpiar todos los datos?"):
            return
        if self.controller:
            self.controller.limpiar_todo()
        self.text_upcs.delete("1.0", "end")
        self.text_fuera.delete("1.0", "end")
        self.tree.delete(*self.tree.get_children())
        for btn in self.botones_transferencias.values():
            btn.destroy()
        self.botones_transferencias.clear()
        for lbl in self.stats.values():
            lbl.configure(text="0")
        self.lbl_ultimo.configure(text="— sin escaneos —")

    def _on_exportar_reporte(self):
        archivo = filedialog.asksaveasfilename(
            defaultextension=".xlsx", filetypes=[("Excel", "*.xlsx")]
        )
        if not archivo:
            return
        wb = Workbook()
        ws = wb.active
        ws.title = "Reporte"
        enc  = Font(bold=True, color="FFFFFF", name="Segoe UI", size=11)
        body = Font(name="Segoe UI", size=11)
        fill_enc = PatternFill(start_color="C62828", end_color="C62828", fill_type="solid")
        borde = Border(**{s: Side(style="thin", color="BDBDBD") for s in ("left","right","top","bottom")})
        centro = Alignment(horizontal="center", vertical="center")
        fills = {
            "Faltante":  PatternFill(start_color="FFEBEE", end_color="FFEBEE", fill_type="solid"),
            "Completo":  PatternFill(start_color="E8F5E9", end_color="E8F5E9", fill_type="solid"),
            "Excedente": PatternFill(start_color="E3F2FD", end_color="E3F2FD", fill_type="solid"),
            "Sobrante":  PatternFill(start_color="FFF8E1", end_color="FFF8E1", fill_type="solid"),
        }
        ws.append(["Código", "Descripción", "Transferencia", "Pistoleada", "Estado"])
        for col in range(1, 6):
            c = ws.cell(row=1, column=col)
            c.font, c.fill, c.alignment, c.border = enc, fill_enc, centro, borde
        for i, fila in enumerate(self.tree.get_children(), 2):
            v = self.tree.item(fila, "values")
            t, p = int(v[2]), int(v[3])
            if t > p:        estado, fill = "Faltante",  fills["Faltante"]
            elif t == p > 0: estado, fill = "Completo",  fills["Completo"]
            elif t == 0:     estado, fill = "Excedente", fills["Excedente"]
            else:            estado, fill = "Sobrante",  fills["Sobrante"]
            ws.append([v[0], v[1], t, p, estado])
            for j in range(1, 6):
                c = ws.cell(row=i, column=j)
                c.font, c.border, c.alignment, c.fill = body, borde, centro, fill
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions
        ws.column_dimensions["B"].width = 40
        for col in ws.columns:
            l = col[0].column_letter
            if l == "B": continue
            ml = max((len(str(c.value)) for c in col if c.value), default=8)
            ws.column_dimensions[l].width = ml + 4
        try:
            wb.save(archivo)
            messagebox.showinfo("Éxito", "Reporte guardado.")
        except Exception as e:
            messagebox.showerror("Error", str(e))

    def _on_ordenar(self):
        grupos = {"azul": [], "amarillo": [], "rojo": [], "verde": []}
        for item in self.tree.get_children():
            v = self.tree.item(item, "values")
            t, p = int(v[2]), int(v[3])
            if t == 0:   grupos["azul"].append(v)
            elif p > t:  grupos["amarillo"].append(v)
            elif p < t:  grupos["rojo"].append(v)
            else:        grupos["verde"].append(v)
        self.tree.delete(*self.tree.get_children())
        for tag in ["azul", "amarillo", "rojo", "verde"]:
            for v in grupos[tag]:
                self.tree.insert("", "end", values=v, tags=(tag,))
        self._actualizar_colores()

    # ──────────────────────────────────────────────
    # Helpers UI
    # ──────────────────────────────────────────────

    def _agregar_boton_transferencia(self, nombre: str):
        btn = ctk.CTkButton(
            self.frame_transferencias,
            text=f"  {nombre}",
            height=36,
            corner_radius=4,
            fg_color=C_RED_TINT,
            hover_color="#FFCDD2",
            text_color=C_PRIMARY,
            font=ctk.CTkFont("Segoe UI", 12, weight="bold"),
            anchor="w",
            command=lambda: self._on_eliminar_transferencia(nombre),
        )
        btn.pack(fill="x", pady=2)
        self.botones_transferencias[nombre] = btn

    def _on_eliminar_transferencia(self, nombre: str):
        if not messagebox.askyesno("Eliminar", f"¿Eliminar '{nombre}'?"):
            return
        if self.controller:
            filas = self.controller.eliminar_transferencia(nombre)
            self._refrescar_treeview(filas)
            self._actualizar_contadores()
        btn = self.botones_transferencias.pop(nombre, None)
        if btn:
            btn.destroy()

    def _refrescar_treeview(self, filas: list):
        self.tree.delete(*self.tree.get_children())
        for fila in filas:
            self.tree.insert("", "end", values=fila)
        self._actualizar_colores()

    def _actualizar_fila_treeview(self, resultado: dict):
        codigo = resultado["codigo"]
        for item in self.tree.get_children():
            if self.tree.item(item, "values")[0] == codigo:
                self.tree.item(item, values=(
                    codigo, resultado["descripcion"],
                    resultado["cantidad_transferencia"],
                    resultado["cantidad_pistoleada"],
                ))
                return
        self.tree.insert("", "end", values=(
            codigo, resultado["descripcion"],
            resultado["cantidad_transferencia"],
            resultado["cantidad_pistoleada"],
        ))

    def _actualizar_contadores(self):
        if not self.controller:
            return
        c = self.controller.get_contadores()
        self.stats["Transferidas"].configure(text=f"{c['transferidas']:.0f}")
        self.stats["Pistoleadas"].configure(text=f"{c['pistoleadas']:.0f}")
        self.stats["Fuera"].configure(text=str(c["fuera"]))
        self.stats["Diferencia"].configure(text=f"{c['diferencia']:.0f}")

    def _actualizar_colores(self):
        for item in self.tree.get_children():
            v = self.tree.item(item, "values")
            t, p = int(v[2]), int(v[3])
            if t == 0:    tag = "azul"
            elif p == t:  tag = "verde"
            elif p > t:   tag = "amarillo"
            else:         tag = "rojo"
            if self.tree.item(item, "tags") != (tag,):
                self.tree.item(item, tags=(tag,))


if __name__ == "__main__":
    app = MainView()
    app.mainloop()